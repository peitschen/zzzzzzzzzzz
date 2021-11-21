from xml.dom import minidom
import openpyxl 
import os
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl import Workbook, load_workbook
from copy import copy
from openpyxl.utils import get_column_letter
import msvcrt
abspath = os.getcwd() 
bakpath = os.getcwd()
rootpath = os.path.abspath('..')  
ret = abspath.replace(rootpath, '', 1)
ret=ret[1:]

def split_list(a_list):
    half = len(a_list)//2
    return a_list[:half]
    half = len(a_list)//2
    return a_list[:half]

def subdir_list(dirname):
    return list(filter(os.path.isdir,map(lambda filename: os.path.join(dirname, filename),os.listdir(dirname))))
def xlsx_sheet_copy(src_path, tag_path, sheet_name):  
    src_workbook = load_workbook(src_path)  
    src_file_sheet = src_workbook[sheet_name]  
    tag_workbook = load_workbook(tag_path) 
    tag_file_sheet = tag_workbook.create_sheet(sheet_name)  
    for row in src_file_sheet:      
        for cell in row:  
            tag_file_sheet[cell.coordinate].value = cell.value
            if cell.has_style: 
                tag_file_sheet[cell.coordinate].font = copy(cell.font)
                tag_file_sheet[cell.coordinate].border = copy(cell.border)
                tag_file_sheet[cell.coordinate].fill = copy(cell.fill)
                tag_file_sheet[cell.coordinate].number_format = copy(cell.number_format)
                tag_file_sheet[cell.coordinate].protection = copy(cell.protection)
                tag_file_sheet[cell.coordinate].alignment = copy(cell.alignment)
    wm = list(zip(src_file_sheet.merged_cells))  
    if len(wm) > 0:  
        for i in range(0, len(wm)):
            cell2 = (
                str(wm[i]).replace("(<MergedCellRange ", "").replace(">,)", "")
            )  
            tag_file_sheet.merge_cells(cell2)
    for i in range(1, src_file_sheet.max_row + 1):
        tag_file_sheet.row_dimensions[i].height = src_file_sheet.row_dimensions[
            i
        ].height
    for i in range(1, src_file_sheet.max_column + 1):
        tag_file_sheet.column_dimensions[
            get_column_letter(i)
        ].width = src_file_sheet.column_dimensions[get_column_letter(i)].width
    tag_workbook.save(tag_path)  
    tag_workbook.close()  
    src_workbook.close()
dirs=subdir_list(abspath)
z=0
while len(dirs) >=1 :
    os.chdir(dirs[0])   
    if os.path.exists('sample.acaml'):
        abspath1 = os.getcwd() 
        rootpath1 = os.path.abspath('..')  
        ret1 = abspath1.replace(rootpath1, '', 1)
        ret1=ret1[1:]
        wb=openpyxl.Workbook()
        ws = wb.active
        ws.title=str(ret1)
        dom = minidom.parse('sample.acaml')
        root = dom.documentElement
        RetentionTimes=[]
        RetentionTime=dom.getElementsByTagName("RetentionTime") 
        for i in range(len(RetentionTime)):
            RetentionTimes.append(RetentionTime[i].getAttribute("val"))
        #del RetentionTimes[0: 2]
        RetentionTimes.insert(0,"时间")
        for i in range(len(RetentionTimes)):
            ws.cell(1,i+1,RetentionTimes[i])
        Areas=[]
        Area=dom.getElementsByTagName("Area") 
        for i in range(len(Area)): 
            Areas.append(Area[i].getAttribute("val"))
        Areas=split_list(Areas)
        Areas.insert(0,"峰面积")
        for i in range(len(Areas)):
            ws.cell(2,i+1,Areas[i])
        Heights=[]
        Height=dom.getElementsByTagName("Height") 
        for i in range(len(Height)):
            Heights.append(Height[i].getAttribute("val"))
        Heights=split_list(Heights)
        Heights.insert(0,"峰高")
        for i in range(len(Heights)):
            ws.cell(3,i+1,Heights[i])
        WidthBases=[]
        WidthBase=dom.getElementsByTagName("WidthBase")  
        for i in range(len(WidthBase)):  
            WidthBases.append(WidthBase[i].getAttribute("val"))
        WidthBases.insert(0,"峰宽")
        for i in range(len(WidthBases)):
            ws.cell(4,i+1,WidthBases[i])
        Symmetrys=[]
        Symmetry=dom.getElementsByTagName("Symmetry")  
        for i in range(len(Symmetry)):  
            Symmetrys.append(Symmetry[i].getAttribute("val"))
        Symmetrys.insert(0,"对称因子")
        for i in range(len(Symmetrys)):
            ws.cell(5,i+1,Symmetrys[i])
    
        AreaPercents=[]
        AreaPercent=dom.getElementsByTagName("AreaPercent") 
        for i in range(len(AreaPercent)):  
            AreaPercents.append(AreaPercent[i].getAttribute("val"))
        AreaPercents.insert(0,"峰面积占比")
        for i in range(len(AreaPercents)):
            ws.cell(6,i+1,AreaPercents[i])


        wb.save(filename=str(ret1)+'.xlsx')
        dirs.pop(0)
        z=z+1
    else:
        dirs.pop(0)

os.chdir(bakpath)
dirs=subdir_list(bakpath)
if z==0:
    print("未成功")
else:
    print("成功创建了%d个工作表"%z)
wball=openpyxl.Workbook()
wball.save(filename='all.xlsx')
z=0
while len(dirs) >=1 :
    os.chdir(dirs[0])
    abspath1 = os.getcwd()
    rootpath1 = os.path.abspath('..') 
    ret1 = abspath1.replace(rootpath1, '', 1)
    ret1=ret1[1:]
    if os.path.exists(str(ret1)+'.xlsx'):
        xlsx_sheet_copy(str(dirs[0]+'\\'+str(ret1)+'.xlsx'),str(bakpath)+'\\'+'all.xlsx',str(ret1))
        dirs.pop(0)
        z=z+1
    else:
        dirs.pop(0)

sheet_name = 'Sheet'
workbook = openpyxl.load_workbook(str(bakpath)+'\\'+'all.xlsx')
worksheet = workbook[sheet_name]
workbook.remove(worksheet)
workbook.save(str(bakpath)+'\\'+'all.xlsx')
if z==0:
    print("未成功")
else:
    print("成功合并了%d个工作表"%z)
print("执行完毕，请按任意键退出")
ord(msvcrt.getch())
